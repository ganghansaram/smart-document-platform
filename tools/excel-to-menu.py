#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to menu.json 변환 스크립트

이 스크립트는 엑셀 파일의 트리 구조를 menu.json 파일로 변환합니다.

===========================================
전제조건 및 엑셀 파일 형식
===========================================

1. 필수 라이브러리 설치:
   pip install openpyxl

2. 엑셀 파일 구조:
   - 첫 번째 행: 헤더 (Level1, Level2, ..., LevelN, URL)
   - 두 번째 행부터: 데이터
   - 레벨 깊이는 자동 감지 (몇 단계든 가능)

3. 컬럼 구성 예시 (레벨 5개 + URL):
   | A (Level1) | B (Level2) | C (Level3) | D (Level4) | E (Level5) | F (URL) |
   |------------|------------|------------|------------|------------|---------|
   | 개발 개요   |            |            |            |            |         |
   |            | 프로그램 소개|            |            |            | contents/dev-overview/introduction.html |
   |            | 개발 히스토리|            |            |            |         |
   |            |            | 초기 연구   |            |            | contents/dev-overview/history/concept.html |
   |            |            | 체계개발    |            |            | contents/dev-overview/history/development.html |
   | 시스템 설계 |            |            |            |            |         |
   ...

4. 규칙:
   - 마지막 컬럼은 항상 URL로 처리
   - 그 외 컬럼들은 레벨 컬럼 (레벨 깊이 자동 감지)
   - 각 행에서 값이 있는 가장 왼쪽 컬럼이 해당 항목의 레벨을 결정
   - 상위 레벨이 비어있으면 이전에 정의된 상위 항목의 하위로 들어감
   - URL 컬럼이 비어있으면 폴더로 처리 (클릭 시 펼침/접힘만 동작)
   - URL 컬럼에 값이 있으면 문서로 처리 (클릭 시 해당 페이지 로드)

5. 사용법:
   python excel-to-menu.py <엑셀파일경로> [출력파일경로]

   예시:
   python excel-to-menu.py menu_tree.xlsx
   python excel-to-menu.py menu_tree.xlsx ../data/menu.json

===========================================
"""

import json
import sys
import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("오류: openpyxl 라이브러리가 필요합니다.")
    print("설치 명령: pip install openpyxl")
    sys.exit(1)


def read_excel_tree(excel_path):
    """
    엑셀 파일에서 트리 데이터를 읽어옵니다.
    레벨 깊이는 자동으로 감지됩니다 (마지막 컬럼 = URL).

    Returns:
        tuple: (rows, level_count)
            - rows: [(level, label, url), ...] 형태의 리스트
            - level_count: 감지된 레벨 깊이
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    # 첫 번째 행(헤더)에서 컬럼 수 감지
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    total_cols = len([c for c in header_row if c is not None])
    level_count = total_cols - 1  # 마지막 컬럼은 URL

    print(f"감지된 레벨 깊이: {level_count}")

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=total_cols, values_only=True), start=2):
        # 레벨 컬럼에서 값 찾기
        level = None
        label = None

        for col_idx in range(level_count):
            cell_value = row[col_idx] if col_idx < len(row) else None
            if cell_value is not None and str(cell_value).strip():
                level = col_idx + 1  # 1-based level
                label = str(cell_value).strip()
                break

        if level is None or label is None:
            continue  # 빈 행 스킵

        # URL (마지막 컬럼)
        url_idx = level_count
        url = row[url_idx] if len(row) > url_idx and row[url_idx] else None
        if url:
            url = str(url).strip()

        rows.append((level, label, url))

    wb.close()
    return rows, level_count


def build_tree(rows):
    """
    플랫 리스트를 계층 트리 구조로 변환합니다.

    Args:
        rows: [(level, label, url), ...] 형태의 리스트

    Returns:
        list: menu.json 형식의 트리 구조
    """
    root = []
    stack = [(0, root)]  # (level, children_list)

    for level, label, url in rows:
        # 현재 레벨보다 깊거나 같은 스택 항목 제거
        while stack and stack[-1][0] >= level:
            stack.pop()

        # 새 항목 생성
        item = {"label": label}
        if url:
            item["url"] = url

        # 부모의 children에 추가
        parent_children = stack[-1][1]
        parent_children.append(item)

        # 현재 항목을 스택에 추가 (나중에 자식이 올 수 있으므로)
        item["children"] = []
        stack.append((level, item["children"]))

    # 빈 children 제거
    def remove_empty_children(items):
        for item in items:
            if "children" in item:
                if item["children"]:
                    remove_empty_children(item["children"])
                else:
                    del item["children"]

    remove_empty_children(root)

    return root


def main():
    # 인자 처리
    if len(sys.argv) < 2:
        print("사용법: python excel-to-menu.py <엑셀파일경로> [출력파일경로]")
        print("예시: python excel-to-menu.py menu_tree.xlsx ../data/menu.json")
        sys.exit(1)

    excel_path = sys.argv[1]

    # 기본 출력 경로: 스크립트 위치 기준 ../data/menu.json
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_output = os.path.join(script_dir, "..", "data", "menu.json")
    output_path = sys.argv[2] if len(sys.argv) > 2 else default_output

    # 엑셀 파일 존재 확인
    if not os.path.exists(excel_path):
        print(f"오류: 엑셀 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)

    print(f"엑셀 파일 읽는 중: {excel_path}")

    # 엑셀에서 데이터 읽기
    rows, level_count = read_excel_tree(excel_path)
    print(f"읽은 항목 수: {len(rows)}")

    # 트리 구조로 변환
    tree = build_tree(rows)

    # JSON 파일로 저장
    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(tree, f, ensure_ascii=False, indent=2)

    print(f"menu.json 생성 완료: {output_path}")
    print(f"총 최상위 메뉴 수: {len(tree)}")


if __name__ == "__main__":
    main()
