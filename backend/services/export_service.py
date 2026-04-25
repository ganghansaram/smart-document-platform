"""
Verify 내보내기 서비스 — Excel(.xlsx) 리포트 생성

프론트엔드 buildExportPayload()가 전송한 JSON을 받아
openpyxl로 Excel 파일을 생성하여 BytesIO 스트림으로 반환한다.
"""
import io
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ── 공통 스타일 ──
_HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_FONT = Font(name="맑은 고딕", size=10)
_BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

_SEVERITY_FILLS = {
    "오류": PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),
    "경고": PatternFill(start_color="FEF3CD", end_color="FEF3CD", fill_type="solid"),
    "제안": PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
}

_DECISION_FILLS = {
    "수락": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    "거절": PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),
    "미처리": PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
}


def _apply_header_style(ws, row, col_count):
    """헤더 행에 스타일을 적용한다."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _apply_body_style(ws, row, col_count, align=None):
    """본문 행에 스타일을 적용한다."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _BODY_FONT
        cell.alignment = align or _BODY_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws, col_count, max_width=50):
    """컬럼 너비를 내용에 맞게 자동 조정한다."""
    for col in range(1, col_count + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    length = max(len(str(cell.value).split("\n")[0]), 4)
                    if length > max_len:
                        max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _write_summary_sheet(ws, data):
    """공통 요약 시트를 작성한다."""
    ws.title = "요약"
    mode_labels = {"compare": "문서 비교", "similarity": "유사도 검사", "verify": "규칙 검증"}
    mode = data.get("mode", "")

    rows = [
        ("검토 모드", mode_labels.get(mode, mode)),
        ("검토일시", data.get("timestamp", "")[:19].replace("T", " ")),
        ("검토자", data.get("reviewer", "")),
    ]

    files = data.get("files", {})
    if files.get("a"):
        rows.append(("문서 A", files["a"]))
    if files.get("b"):
        rows.append(("문서 B", files["b"]))

    # 모드별 핵심 수치
    if mode == "compare":
        s = data.get("summary", {})
        rows.append(("총 변경", f"{s.get('total', 0)}건"))
        rows.append(("수락", f"{s.get('accepted', 0)}건"))
        rows.append(("거절", f"{s.get('rejected', 0)}건"))
        rows.append(("미처리", f"{s.get('undecided', 0)}건"))
    elif mode == "similarity":
        # Plan-45 v3: 유사도 모드는 Excel export UI 폐기 (PDF/HTML 만 노출), 본 분기는 dead path 이지만
        # 다른 호출 경로 (백엔드 직접 호출 등) 대비 4 카테고리 어휘로 갱신
        rows.append(("유사율", f"{data.get('score', 0)}%"))
        verdict_display = data.get("verdict_label") or data.get("verdict_legacy") or data.get("verdict", "")
        rows.append(("판정", verdict_display))
        rows.append(("전체 문장", f"{data.get('total_sentences', 0)}문장"))
        # tiers (Plan-38 잔존) — v3에서는 카테고리 카운트로 대체. 백엔드 호환을 위해 유지.
        tiers = data.get("tiers", {})
        if tiers.get("substantive") is not None:
            rows.append(("동일·거의 동일", f"{tiers['substantive']}%"))
            rows.append(("의역", f"{tiers.get('derived', 0)}%"))
            rows.append(("공통 정형구문", f"{tiers.get('boilerplate', 0)}%"))
            if tiers.get("excluded") is not None:
                rows.append(("제외 문장", f"{tiers['excluded']}%"))
        # Phase 2: sources
        sources = data.get("sources") or []
        if sources:
            for i, src in enumerate(sources, 1):
                rows.append((f"출처 [{src.get('id', i)}]", f"{src.get('name', '')} — {src.get('match_pct', 0)}% ({src.get('matched_sents', 0)}문장 / {src.get('matched_words', 0)}단어)"))
    elif mode == "verify":
        rows.append(("점수", f"{data.get('score', 0)}점"))
        rows.append(("등급", data.get("grade", "")))
        s = data.get("summary", {})
        rows.append(("오류", f"{s.get('error', 0)}건"))
        rows.append(("경고", f"{s.get('warning', 0)}건"))
        rows.append(("제안", f"{s.get('suggestion', 0)}건"))

    # 타이틀
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value=f"{mode_labels.get(mode, mode)} 리포트")
    title_cell.font = Font(name="맑은 고딕", bold=True, size=14, color="2C5282")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(name="맑은 고딕", bold=True, size=10)
        ws.cell(row=i, column=2, value=value).font = _BODY_FONT
        ws.cell(row=i, column=1).border = _THIN_BORDER
        ws.cell(row=i, column=2).border = _THIN_BORDER

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 40


def _write_compare_sheet(ws, data):
    """비교 모드 변경점 시트를 작성한다."""
    ws.title = "변경점"
    headers = ["#", "판정", "AI분류", "원문", "수정문"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    changes = data.get("changes", [])
    for i, c in enumerate(changes, start=2):
        ws.cell(row=i, column=1, value=c.get("index", i - 1))
        dec_cell = ws.cell(row=i, column=2, value=c.get("decision", ""))
        ws.cell(row=i, column=3, value=c.get("ai_tag", ""))
        ws.cell(row=i, column=4, value=c.get("old_text", "")[:500])
        ws.cell(row=i, column=5, value=c.get("new_text", "")[:500])
        _apply_body_style(ws, i, len(headers))
        ws.cell(row=i, column=1).alignment = _CENTER_ALIGN
        ws.cell(row=i, column=2).alignment = _CENTER_ALIGN
        ws.cell(row=i, column=3).alignment = _CENTER_ALIGN
        # 판정별 배경색
        fill = _DECISION_FILLS.get(c.get("decision", ""))
        if fill:
            dec_cell.fill = fill

    _auto_width(ws, len(headers))
    # 텍스트 열은 고정 너비
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 50


def _write_similarity_sheet(ws, data):
    """유사도 모드 매칭 시트를 작성한다."""
    ws.title = "매칭 목록"
    headers = ["#", "유형", "유사율", "분석방법", "대상문장", "참조문장"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    matches = data.get("matches", [])
    for i, m in enumerate(matches, start=2):
        ws.cell(row=i, column=1, value=m.get("index", i - 1))
        ws.cell(row=i, column=2, value=m.get("type", ""))
        ws.cell(row=i, column=3, value=m.get("similarity", ""))
        ws.cell(row=i, column=4, value=m.get("method", ""))
        ws.cell(row=i, column=5, value=m.get("target_text", "")[:500])
        ws.cell(row=i, column=6, value=m.get("ref_text", "")[:500])
        _apply_body_style(ws, i, len(headers))
        ws.cell(row=i, column=1).alignment = _CENTER_ALIGN
        ws.cell(row=i, column=2).alignment = _CENTER_ALIGN
        ws.cell(row=i, column=3).alignment = _CENTER_ALIGN
        ws.cell(row=i, column=4).alignment = _CENTER_ALIGN

    _auto_width(ws, len(headers))
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 50


def _write_verify_sheet(ws, data):
    """규칙 검증 모드 이슈 시트를 작성한다."""
    ws.title = "이슈 목록"
    headers = ["#", "심각도", "규칙", "위치", "내용"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    issues = data.get("issues", [])
    for i, iss in enumerate(issues, start=2):
        ws.cell(row=i, column=1, value=iss.get("index", i - 1))
        sev_cell = ws.cell(row=i, column=2, value=iss.get("severity", ""))
        ws.cell(row=i, column=3, value=iss.get("rule", ""))
        ws.cell(row=i, column=4, value=iss.get("location", ""))
        ws.cell(row=i, column=5, value=iss.get("message", ""))
        _apply_body_style(ws, i, len(headers))
        ws.cell(row=i, column=1).alignment = _CENTER_ALIGN
        ws.cell(row=i, column=2).alignment = _CENTER_ALIGN
        ws.cell(row=i, column=4).alignment = _CENTER_ALIGN
        # 심각도별 배경색
        fill = _SEVERITY_FILLS.get(iss.get("severity", ""))
        if fill:
            sev_cell.fill = fill

    _auto_width(ws, len(headers))
    ws.column_dimensions["E"].width = 60


def _write_intelligence_sheet(ws, data):
    """인텔리전스 분석 시트 (verify 모드 전용)."""
    ws.title = "문서 분석"
    row = 1

    # 1. 문서 구조
    structure = data.get("structure")
    if structure:
        ws.cell(row=row, column=1, value="문서 구조").font = Font(name="맑은 고딕", bold=True, size=12, color="2C5282")
        row += 1
        stats = [
            ("단락", structure.get("total_paragraphs", 0)),
            ("문장", structure.get("total_sentences", 0)),
            ("섹션", len(structure.get("headings", []))),
            ("표", len(structure.get("tables", []))),
            ("그림", len(structure.get("figures", []))),
        ]
        for label, val in stats:
            ws.cell(row=row, column=1, value=label).font = Font(name="맑은 고딕", bold=True, size=10)
            ws.cell(row=row, column=2, value=val).font = _BODY_FONT
            ws.cell(row=row, column=1).border = _THIN_BORDER
            ws.cell(row=row, column=2).border = _THIN_BORDER
            row += 1

        headings = structure.get("headings", [])
        sections = structure.get("sections", [])
        if headings:
            row += 1
            for col, h in enumerate(["번호", "제목", "단락 수"], 1):
                ws.cell(row=row, column=col, value=h)
            _apply_header_style(ws, row, 3)
            row += 1
            for hi, hd in enumerate(headings):
                ws.cell(row=row, column=1, value=f"{hd['number']}.").font = _BODY_FONT
                ws.cell(row=row, column=1).alignment = _CENTER_ALIGN
                ws.cell(row=row, column=2, value=hd.get("text", "")).font = _BODY_FONT
                pc = sections[hi]["paragraph_count"] if hi < len(sections) else ""
                ws.cell(row=row, column=3, value=pc).font = _BODY_FONT
                ws.cell(row=row, column=3).alignment = _CENTER_ALIGN
                _apply_body_style(ws, row, 3)
                row += 1

    # 2. 요구사항 분류
    requirements = data.get("requirements")
    if requirements and requirements.get("counts"):
        row += 1
        ws.cell(row=row, column=1, value="요구사항 분류").font = Font(name="맑은 고딕", bold=True, size=12, color="2C5282")
        row += 1
        for col, h in enumerate(["분류", "건수"], 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, 2)
        row += 1
        labels = {"shall": "필수 (shall)", "should": "권고 (should)", "may": "허용 (may)",
                  "test_condition": "시험 조건", "information": "정보 서술"}
        for key in ["shall", "should", "may", "test_condition", "information"]:
            count = requirements["counts"].get(key, 0)
            if count > 0:
                ws.cell(row=row, column=1, value=labels[key]).font = _BODY_FONT
                ws.cell(row=row, column=2, value=count).font = _BODY_FONT
                ws.cell(row=row, column=2).alignment = _CENTER_ALIGN
                _apply_body_style(ws, row, 2)
                row += 1

    # 3. 용어/규격
    terms = data.get("terms")
    if terms:
        specs = terms.get("specifications", [])
        abbrs = terms.get("abbreviations", [])
        units = terms.get("units", [])

        if specs:
            row += 1
            ws.cell(row=row, column=1, value="규격 번호").font = Font(name="맑은 고딕", bold=True, size=12, color="2C5282")
            row += 1
            for col, h in enumerate(["규격", "출현 횟수"], 1):
                ws.cell(row=row, column=col, value=h)
            _apply_header_style(ws, row, 2)
            row += 1
            for s in specs:
                ws.cell(row=row, column=1, value=s["id"]).font = _BODY_FONT
                ws.cell(row=row, column=2, value=s["count"]).font = _BODY_FONT
                ws.cell(row=row, column=2).alignment = _CENTER_ALIGN
                _apply_body_style(ws, row, 2)
                row += 1

        if abbrs:
            row += 1
            ws.cell(row=row, column=1, value="약어").font = Font(name="맑은 고딕", bold=True, size=12, color="2C5282")
            row += 1
            for col, h in enumerate(["약어", "정의", "출현 횟수"], 1):
                ws.cell(row=row, column=col, value=h)
            _apply_header_style(ws, row, 3)
            row += 1
            for a in abbrs:
                ws.cell(row=row, column=1, value=a["abbr"]).font = Font(name="맑은 고딕", bold=True, size=10)
                ws.cell(row=row, column=2, value=a.get("definition", "")).font = _BODY_FONT
                ws.cell(row=row, column=3, value=a["count"]).font = _BODY_FONT
                ws.cell(row=row, column=3).alignment = _CENTER_ALIGN
                _apply_body_style(ws, row, 3)
                row += 1

        if units:
            row += 1
            ws.cell(row=row, column=1, value="단위").font = Font(name="맑은 고딕", bold=True, size=12, color="2C5282")
            row += 1
            for col, h in enumerate(["단위", "출현 횟수", "SI 준수"], 1):
                ws.cell(row=row, column=col, value=h)
            _apply_header_style(ws, row, 3)
            row += 1
            for u in units:
                ws.cell(row=row, column=1, value=u["unit"]).font = _BODY_FONT
                ws.cell(row=row, column=2, value=u["count"]).font = _BODY_FONT
                ws.cell(row=row, column=2).alignment = _CENTER_ALIGN
                si_cell = ws.cell(row=row, column=3, value="✓" if u.get("si_compliant") else "⚠")
                si_cell.font = _BODY_FONT
                si_cell.alignment = _CENTER_ALIGN
                _apply_body_style(ws, row, 3)
                row += 1

    _auto_width(ws, 3)
    ws.column_dimensions["B"].width = 40


def _write_similarity_criteria_sheet(ws, help_data: dict | None = None):
    """유사도 검사 기준 시트 (Plan-38 §5.6 — L4 부록 자동 첨부).

    SSOT JSON에서 산식·신호등·라벨 정의·면책을 로드. 없으면 인라인 폴백.
    """
    ws.title = "검사 기준"

    # 폴백 데이터 (help_data 미제공 시)
    fallback_bands = [
        {"color": "Blue",   "range_min": 0,  "range_max": 0,   "label": "매칭 없음",   "meaning": "완전 독창"},
        {"color": "Green",  "range_min": 1,  "range_max": 24,  "label": "양호",        "meaning": "정상 인용·정형구문 수준"},
        {"color": "Yellow", "range_min": 25, "range_max": 49,  "label": "검토 필요",   "meaning": "출처 표기·재서술 점검"},
        {"color": "Orange", "range_min": 50, "range_max": 74,  "label": "상당량 매칭", "meaning": "광범위한 재작성 권고"},
        {"color": "Red",    "range_min": 75, "range_max": 100, "label": "위험",        "meaning": "대부분 타 출처와 동일"},
    ]
    # Plan-45 v3: SSOT 정상 시 자동 동기, fallback도 v3 어휘
    fallback_labels = {
        "identical":   {"ko_long": "동일 (직접 차용)",        "lex": "매우 높음", "sem": "—",   "short": "단어까지 거의 그대로",         "threshold": "fp ≥ 85%"},
        "near_copy":   {"ko_long": "거의 동일 (단어 변형)",   "lex": "높음", "sem": "높음", "short": "단어 일부만 바꿈",             "threshold": "fp 40~85% + sem ≥ 0.85"},
        "paraphrase":  {"ko_long": "의역 (재서술)",           "lex": "낮음", "sem": "높음", "short": "단어 다른데 같은 말",          "threshold": "fp < 40% + sem ≥ 0.75"},
        "translation": {"ko_long": "의역 (다른 언어 번역)",    "lex": "거의 0", "sem": "높음", "short": "한↔영 등 언어 차이 — 의역 통합", "threshold": "fp < 10% + sem ≥ 0.75 + 다른 스크립트"},
        "low_sim":     {"ko_long": "약한 유사 (참고용)",       "lex": "낮음", "sem": "중간", "short": "단정 어려운 약한 관련성",       "threshold": "sem 0.65~0.75"},
        "boilerplate": {"ko_long": "공통 정형구문 (제외)",    "lex": "—",   "sem": "—",   "short": "업계 표준 문구",               "threshold": "phrase coverage ≥ 50%"},
    }
    fallback_formula = "유사율 = (동일 + 거의 동일 + 의역) / (전체 문장 - 제외 문장) × 100"
    fallback_disclaimer = "유사도 ≠ 표절 — 검토자의 판단이 최종 (Crossref Similarity Check 가이드)"

    bands = (help_data or {}).get("verdict_bands") or fallback_bands
    labels = (help_data or {}).get("labels") or fallback_labels
    formula = ((help_data or {}).get("score_formula") or {}).get("equation") or fallback_formula
    disclaimer = ((help_data or {}).get("disclaimer") or {}).get("main") or fallback_disclaimer

    row = 1
    # 타이틀
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    title = ws.cell(row=row, column=1, value="유사도 검사 기준 (Reference)")
    title.font = Font(name="맑은 고딕", bold=True, size=14, color="2C5282")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 2

    # 산식
    sec_h = ws.cell(row=row, column=1, value="산식")
    sec_h.font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    sec_h.fill = _HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    eq_cell = ws.cell(row=row, column=1, value=formula)
    eq_cell.font = Font(name="Consolas", size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # 5단계 신호등
    sec_h = ws.cell(row=row, column=1, value="5단계 신호등")
    sec_h.font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    sec_h.fill = _HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    headers = ["색상", "구간(%)", "판정", "의미"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid")
        c.alignment = _HEADER_ALIGN
        c.border = _THIN_BORDER
    row += 1
    band_colors = {"blue": "EFF6FF", "green": "F0FDF4", "yellow": "FEFCE8", "orange": "FFF7ED", "red": "FEF2F2"}
    for b in bands:
        color_key = (b.get("color") or "").lower()
        bg = band_colors.get(color_key, "FFFFFF")
        rng_str = f"{b.get('range_min', 0)}" if b.get("range_min") == b.get("range_max") else f"{b.get('range_min', 0)}~{b.get('range_max', 0)}"
        cells = [
            (b.get("color") or "").upper(),
            rng_str,
            b.get("label", ""),
            b.get("meaning", ""),
        ]
        for col, val in enumerate(cells, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = _BODY_FONT
            c.alignment = _BODY_ALIGN
            c.border = _THIN_BORDER
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        row += 1
    row += 1

    # 매칭 유형 (6종)
    sec_h = ws.cell(row=row, column=1, value="매칭 유형 (6종)")
    sec_h.font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    sec_h.fill = _HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    headers2 = ["유형", "어휘", "의미", "설명 / 임계값"]
    for col, h in enumerate(headers2, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid")
        c.alignment = _HEADER_ALIGN
        c.border = _THIN_BORDER
    row += 1
    for key in ("identical", "near_copy", "paraphrase", "translation", "low_sim", "boilerplate"):
        lbl = labels.get(key, {})
        if not lbl:
            continue
        desc = f"{lbl.get('short', '')} · {lbl.get('threshold', '')}"
        cells = [lbl.get("ko_long") or lbl.get("ko") or key, lbl.get("lex", "—"), lbl.get("sem", "—"), desc]
        for col, val in enumerate(cells, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = _BODY_FONT
            c.alignment = _BODY_ALIGN
            c.border = _THIN_BORDER
        row += 1
    row += 1

    # 면책
    sec_h = ws.cell(row=row, column=1, value="면책 사항")
    sec_h.font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    sec_h.fill = _HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    dc_cell = ws.cell(row=row, column=1, value=disclaimer)
    dc_cell.font = Font(name="맑은 고딕", italic=True, size=10, color="92400E")
    dc_cell.fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    dc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 36

    # 컬럼 너비
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 50


def _load_similarity_help() -> dict | None:
    """SSOT JSON 로드 (없으면 None — 폴백 사용)."""
    fp = Path(__file__).parent.parent.parent / "data" / "help" / "similarity-help.json"
    if not fp.exists():
        return None
    try:
        import json
        with open(fp, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def build_excel_report(data: dict) -> io.BytesIO:
    """프론트엔드 페이로드로 Excel 리포트를 생성하여 BytesIO 스트림을 반환한다."""
    wb = Workbook()
    mode = data.get("mode", "")

    # 시트 1: 요약
    ws_summary = wb.active
    _write_summary_sheet(ws_summary, data)

    # 시트 2: 모드별 상세
    if mode == "compare":
        ws_detail = wb.create_sheet()
        _write_compare_sheet(ws_detail, data)
    elif mode == "similarity":
        ws_detail = wb.create_sheet()
        _write_similarity_sheet(ws_detail, data)
        # Phase 4: 시트 3 — 검사 기준 (L4 도움말 자동 첨부)
        ws_criteria = wb.create_sheet()
        _write_similarity_criteria_sheet(ws_criteria, _load_similarity_help())
    elif mode == "verify":
        ws_detail = wb.create_sheet()
        _write_verify_sheet(ws_detail, data)
        # 시트 3: 인텔리전스 분석 (verify 전용)
        if data.get("structure") or data.get("requirements") or data.get("terms"):
            ws_intel = wb.create_sheet()
            _write_intelligence_sheet(ws_intel, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
